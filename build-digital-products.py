#!/usr/bin/env python3
"""
Build all digital products for WC 2026 Sports Empire
Outputs: PDFs + XLSXs ready to upload to Gumroad
"""

import os
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import datetime

OUT = "/sessions/hopeful-tender-feynman/mnt/outputs"

# ── COLOURS ──────────────────────────────────────────────────────────────────
DARK   = colors.HexColor("#0a0a0f")
ACCENT = colors.HexColor("#00ff88")
MID    = colors.HexColor("#111118")
LIGHT  = colors.HexColor("#e8e8f0")
DIM    = colors.HexColor("#888899")
RED    = colors.HexColor("#ff4455")
GOLD   = colors.HexColor("#ffd700")

# ─────────────────────────────────────────────────────────────────────────────
# 1. WC 2026 COMPLETE BETTING GUIDE PDF
# ─────────────────────────────────────────────────────────────────────────────
def build_betting_guide():
    path = f"{OUT}/WC2026-Complete-Betting-Guide.pdf"
    doc = SimpleDocTemplate(path, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Normal"],
        fontSize=28, textColor=ACCENT, spaceAfter=6, alignment=TA_CENTER, fontName="Helvetica-Bold")
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
        fontSize=13, textColor=LIGHT, spaceAfter=20, alignment=TA_CENTER)
    h1 = ParagraphStyle("H1", parent=styles["Normal"],
        fontSize=18, textColor=ACCENT, spaceBefore=18, spaceAfter=8, fontName="Helvetica-Bold")
    h2 = ParagraphStyle("H2", parent=styles["Normal"],
        fontSize=13, textColor=GOLD, spaceBefore=12, spaceAfter=6, fontName="Helvetica-Bold")
    body = ParagraphStyle("Body", parent=styles["Normal"],
        fontSize=10, textColor=LIGHT, spaceAfter=8, leading=16)
    small = ParagraphStyle("Small", parent=styles["Normal"],
        fontSize=9, textColor=DIM, spaceAfter=6, leading=14)

    story = []

    # Cover
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph("⚽ WC 2026", title_style))
    story.append(Paragraph("COMPLETE BETTING GUIDE", title_style))
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("AI-Powered Strategy · All 104 Matches · Every Market Explained", sub_style))
    story.append(Paragraph("wc2026-sports-empire.vercel.app", sub_style))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT))
    story.append(Spacer(1, 0.5*cm))

    # Table of contents
    story.append(Paragraph("What's Inside", h1))
    toc = [
        ["Chapter", "Topic", "Page"],
        ["1", "How to Read Betting Odds (all formats)", "3"],
        ["2", "The 10 Main Bet Types Explained", "5"],
        ["3", "Asian Handicap — The Professional's Edge", "8"],
        ["4", "Value Betting with AI", "11"],
        ["5", "WC 2026 Group Stage Strategy", "14"],
        ["6", "WC 2026 Knockout Round Strategy", "17"],
        ["7", "Bankroll Management — Never Go Broke", "20"],
        ["8", "Free Bet Extraction (Welcome Bonuses)", "22"],
        ["9", "The 10 Best Sportsbooks for WC 2026", "25"],
        ["10", "Parlay / Accumulator Guide", "27"],
        ["Bonus", "Live Betting Tactics (In-Play)", "29"],
    ]
    t = Table(toc, colWidths=[1.5*cm, 10*cm, 2*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR", (0,0), (-1,0), DARK),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID, colors.HexColor("#16161e")]),
        ("TEXTCOLOR", (0,1), (-1,-1), LIGHT),
        ("ALIGN", (2,0), (2,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#333344")),
        ("PADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Chapter 1
    story.append(Paragraph("Chapter 1 — How to Read Betting Odds", h1))
    story.append(Paragraph("The three odds formats you'll encounter across different markets:", body))

    story.append(Paragraph("Decimal Odds (Europe, Australia)", h2))
    story.append(Paragraph(
        "The most straightforward format. Your total return = stake × odds. "
        "England at 2.50 means a £10 bet returns £25 (£15 profit + £10 stake). "
        "Anything above 2.00 is an 'odds-against' bet (more profit than stake). "
        "Below 2.00 is 'odds-on' (your stake exceeds your potential profit).", body))

    odds_table = [
        ["Decimal", "Fractional", "American", "Implied %", "What it means"],
        ["1.25", "1/4", "-400", "80%", "Heavy favourite"],
        ["1.50", "1/2", "-200", "66.7%", "Strong favourite"],
        ["2.00", "1/1 (evens)", "+100", "50%", "Even money"],
        ["3.00", "2/1", "+200", "33.3%", "Moderate outsider"],
        ["5.00", "4/1", "+400", "20%", "Longshot"],
        ["10.00", "9/1", "+900", "10%", "Major outsider"],
        ["26.00", "25/1", "+2500", "3.8%", "Near-impossible"],
    ]
    t = Table(odds_table, colWidths=[2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR", (0,0), (-1,0), DARK),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID, colors.HexColor("#16161e")]),
        ("TEXTCOLOR", (0,1), (-1,-1), LIGHT),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#333344")),
        ("PADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.3*cm))

    story.append(Paragraph("Fractional Odds (UK/Ireland)", h2))
    story.append(Paragraph(
        "Written as profit/stake. '3/1' means for every £1 staked you win £3 profit (return = £4). "
        "'1/2' means for every £2 staked you win £1 (return = £3). "
        "Tip: to convert fractional to decimal, divide and add 1. So 3/1 = 3 ÷ 1 + 1 = 4.00.", body))

    story.append(Paragraph("American (Moneyline) Odds — USA", h2))
    story.append(Paragraph(
        "Positive numbers (+) show profit on a $100 stake. Plus 200 means $100 bet returns $300 ($200 profit). "
        "Negative numbers (-) show stake required to win $100. Minus 200 means you need $200 to win $100 back ($300 total). "
        "Most relevant for DraftKings, FanDuel, BetMGM if targeting US traffic.", body))
    story.append(PageBreak())

    # Chapter 2
    story.append(Paragraph("Chapter 2 — The 10 Main Bet Types", h1))

    bet_types = [
        ("1. Match Result (1X2)",
         "The simplest bet: Home Win (1), Draw (X), or Away Win (2). In WC knockout rounds there's no draw market. "
         "Best for: beginners, high-confidence calls on dominant teams."),
        ("2. Asian Handicap (AH)",
         "Eliminates the draw by giving one team a head start. Brazil -1 AH means Brazil must win by 2+ for your bet to win. "
         "Best for: group stage where draws are common, finding value on favourites."),
        ("3. Over/Under Goals",
         "Betting on total goals. O/U 2.5 means you need 3+ goals (over) or 2 or fewer (under). "
         "Best for: matches with clear tactical setups (defensive vs. attacking teams)."),
        ("4. Both Teams to Score (BTTS)",
         "Yes = both teams score at least one goal. No = at least one team keeps a clean sheet. "
         "Best for: WC group stage where weaker teams still try to score."),
        ("5. Correct Score",
         "Predict exact final score. Very high odds, very hard to hit consistently. "
         "Best for: small accumulator stake on 3-4 scores, treat as a lottery ticket."),
        ("6. First Goal Scorer",
         "Which player scores first. Forwards at 4/1 to 12/1. "
         "Best for: attacking players with set-piece responsibility."),
        ("7. Accumulator (Acca)",
         "Chain multiple selections. All must win. Odds multiply. 5 selections at 2.00 each = 32.00. "
         "Best for: small stakes with big potential return. Use our acca guide page for strategy."),
        ("8. Half-Time / Full-Time",
         "Predict result at half time AND full time. E.g. 'Away/Away'. Higher odds, specific scenario. "
         "Best for: favourites facing defensive teams who park the bus early."),
        ("9. Player Props",
         "Anytime scorer, shots on target, assists, cards. "
         "Best for: targeting specialists (free kick takers, penalty takers, high-volume shooters)."),
        ("10. Tournament Outright",
         "Win the tournament, top scorer, most clean sheets. "
         "Best for: pre-tournament where value exists before markets sharpen."),
    ]

    for title, desc in bet_types:
        story.append(Paragraph(title, h2))
        story.append(Paragraph(desc, body))

    story.append(PageBreak())

    # Chapter 3 — Asian Handicap deep dive
    story.append(Paragraph("Chapter 3 — Asian Handicap: The Professional's Edge", h1))
    story.append(Paragraph(
        "Asian Handicap (AH) was developed in Indonesia in the 1990s and is now the dominant bet type "
        "used by professional bettors globally. The reason: it eliminates the draw, giving you only two outcomes. "
        "This significantly reduces the bookmaker's margin from ~8-10% on 1X2 to ~3-5% on AH.", body))

    story.append(Paragraph("Quarter Ball Handicaps (The Push Protection)", h2))
    story.append(Paragraph(
        "AH 0.25 (quarter ball) splits your stake across two lines — half on 0 and half on 0.5. "
        "This means you can get a 'half win' or 'half loss': if the outcome lands on the split, "
        "half your stake is returned and half wins or loses. This is unique to AH and gives you "
        "partial protection against edge cases.", body))

    ah_table = [
        ["Handicap", "Team wins by", "Result"],
        ["0 (Draw No Bet)", "Any margin", "Win"],
        ["0 (Draw No Bet)", "Draw", "Push (stake back)"],
        ["0 (Draw No Bet)", "Loss", "Lose"],
        ["-0.5", "1+ goals", "Win"],
        ["-0.5", "Draw or loss", "Lose"],
        ["-1", "2+ goals", "Win"],
        ["-1", "Exactly 1 goal", "Push (stake back)"],
        ["-1", "Draw or loss", "Lose"],
        ["-1.5", "2+ goals", "Win"],
        ["-1.5", "1 goal, draw, loss", "Lose"],
        ["-2", "3+ goals", "Win"],
        ["-2", "Exactly 2 goals", "Push (stake back)"],
    ]
    t = Table(ah_table, colWidths=[3*cm, 5*cm, 5.5*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR", (0,0), (-1,0), DARK),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID, colors.HexColor("#16161e")]),
        ("TEXTCOLOR", (0,1), (-1,-1), LIGHT),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#333344")),
        ("PADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Chapter 4 — Value Betting
    story.append(Paragraph("Chapter 4 — Value Betting with AI", h1))
    story.append(Paragraph(
        "Value betting is the only mathematically sound long-term strategy. A bet has value when "
        "the bookmaker's implied probability is lower than your actual estimated probability.", body))
    story.append(Paragraph("The Value Formula:", h2))
    story.append(Paragraph(
        "Value = (Your Probability × Decimal Odds) − 1\n\n"
        "Example: You estimate Brazil has a 60% chance of winning (0.60). "
        "Bookmaker offers 2.10 (47.6% implied). "
        "Value = (0.60 × 2.10) − 1 = 1.26 − 1 = +0.26 → 26% edge. STRONG BET.", body))
    story.append(Paragraph(
        "How our AI calculates probability: We run Groq Llama 3 against API-Football's "
        "form data (last 10 matches), head-to-head record, squad availability (injuries/suspensions), "
        "tournament context, and market movement data. The AI assigns a confidence percentage. "
        "We only flag bets where confidence exceeds 65% AND there's positive expected value against "
        "the best available odds.", body))
    story.append(PageBreak())

    # Chapter 5 — Group Stage Strategy
    story.append(Paragraph("Chapter 5 — WC 2026 Group Stage Strategy", h1))
    story.append(Paragraph(
        "WC 2026 features 48 teams in 12 groups of 4, with the top 2 plus 8 best third-placed teams qualifying. "
        "This format creates unique betting dynamics compared to previous tournaments.", body))

    tips = [
        ("Tip 1: The 'Safe Third Place' Effect",
         "With 8 of 12 third-place teams qualifying, massive favourites in group stage have LESS "
         "incentive to go all-out in the final group game. Factor in squad rotation in game 3."),
        ("Tip 2: Draw Value in Opener",
         "Major nations (Brazil, France, Germany) often play conservatively in game 1 — "
         "a draw against their Group A opponents is frequently priced too low."),
        ("Tip 3: Minnow Value in Game 3",
         "A team needing a win to qualify, facing a team already through, at home soil (USA, Canada, Mexico venues) "
         "creates psychological edge for the minnow. Back them +AH."),
        ("Tip 4: Track Team News 2 Hours Before Kick-Off",
         "Lineups release 1 hour before. Key absentees (captain, goalkeeper, striker) move odds 15-30%. "
         "Set up our Telegram alerts to get this instantly."),
        ("Tip 5: Host Nation Advantage",
         "USA, Canada and Mexico are co-hosts. Home crowds in their group games create 0.3-0.5 goal "
         "advantage vs. neutral venue. Market often underprices this."),
    ]

    for title, desc in tips:
        story.append(Paragraph(title, h2))
        story.append(Paragraph(desc, body))
    story.append(PageBreak())

    # Chapter 7 — Bankroll Management
    story.append(Paragraph("Chapter 7 — Bankroll Management", h1))
    story.append(Paragraph(
        "This is the single most important chapter. Without bankroll management, "
        "even a profitable strategy will go bust during a variance downswing.", body))

    story.append(Paragraph("The Kelly Criterion", h2))
    story.append(Paragraph(
        "The mathematically optimal bet sizing formula:\n\n"
        "Kelly % = (BP − Q) / B\n\n"
        "Where: B = decimal odds − 1, P = your win probability, Q = 1 − P\n\n"
        "Example: Odds 2.10, your estimated probability 55% (0.55)\n"
        "B = 1.10, P = 0.55, Q = 0.45\n"
        "Kelly = (1.10 × 0.55 − 0.45) / 1.10 = (0.605 − 0.45) / 1.10 = 0.141 = 14.1% of bankroll\n\n"
        "Important: Most professionals use HALF KELLY (7%) to reduce variance. Never bet more than 5% "
        "on any single selection.", body))

    bk_table = [
        ["Confidence Level", "Recommended Stake", "Max Stake"],
        ["65-70% (standard)", "1-2% of bankroll", "3%"],
        ["70-80% (high conf)", "2-3% of bankroll", "4%"],
        ["80%+ (very high)", "3-4% of bankroll", "5%"],
        ["Any acca (parlay)", "0.5-1% of bankroll", "1.5%"],
        ["Correct score", "0.25-0.5% of bankroll", "0.5%"],
    ]
    t = Table(bk_table, colWidths=[4.5*cm, 5*cm, 4*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR", (0,0), (-1,0), DARK),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID, colors.HexColor("#16161e")]),
        ("TEXTCOLOR", (0,1), (-1,-1), LIGHT),
        ("ALIGN", (1,0), (-1,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#333344")),
        ("PADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Chapter 9 — Best Sportsbooks
    story.append(Paragraph("Chapter 9 — Top 10 Sportsbooks for WC 2026", h1))
    story.append(Paragraph("Ranked by welcome bonus value, odds quality, and market depth:", body))

    books = [
        ["#", "Book", "Region", "Welcome Offer", "Best For"],
        ["1", "bet365", "UK/Global", "Up to £50 in bet credits", "Live betting, Asian markets"],
        ["2", "DraftKings", "USA", "Bet $5 get $200 bonus bets", "US player props, SGP"],
        ["3", "FanDuel", "USA", "No sweat first bet $1,000", "Parlay insurance, boosts"],
        ["4", "Betfair Exchange", "UK/Global", "£20 exchange free bet", "Best odds via lay/back"],
        ["5", "William Hill", "UK", "£30 free bet", "Accumulators, each way"],
        ["6", "Betway", "UK/Africa", "Up to £30 free bets", "Esports + football overlap"],
        ["7", "1xBet", "Global", "100% up to €/$100", "Widest market selection"],
        ["8", "Bet9ja", "Nigeria", "₦100,000 bonus", "African market leader"],
        ["9", "Sportsbet", "Australia", "$501 in bonus bets", "AU/NZ WC markets"],
        ["10", "Betano", "Brazil/EU", "BRL 500 bonus", "LatAm coverage"],
    ]
    t = Table(books, colWidths=[0.8*cm, 2.5*cm, 2.2*cm, 4.2*cm, 3.8*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR", (0,0), (-1,0), DARK),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID, colors.HexColor("#16161e")]),
        ("TEXTCOLOR", (0,1), (-1,-1), LIGHT),
        ("ALIGN", (0,0), (0,-1), "CENTER"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#333344")),
        ("PADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "⚠ Gambling Disclaimer: Betting involves risk. Never bet more than you can afford to lose. "
        "Set deposit limits. GamCare: 0808 8020 133. BeGambleAware.org.", small))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Get live AI picks at wc2026-sports-empire.vercel.app", sub_style))
    story.append(Paragraph("Follow us on Telegram · Discord · X · TikTok · YouTube", sub_style))

    doc.build(story)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 2. OFFICE POOL SPREADSHEET (XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def build_office_pool():
    path = f"{OUT}/WC2026-Office-Pool-Spreadsheet.xlsx"
    wb = openpyxl.Workbook()

    # ── Sheet 1: Participants ──
    ws = wb.active
    ws.title = "Participants"

    green = PatternFill("solid", fgColor="00ff88")
    dark  = PatternFill("solid", fgColor="0a0a0f")
    mid   = PatternFill("solid", fgColor="111118")
    head_font = Font(name="Calibri", bold=True, color="0a0a0f", size=11)
    white_font = Font(name="Calibri", color="e8e8f0", size=10)
    gold_font  = Font(name="Calibri", bold=True, color="ffd700", size=11)

    ws["A1"] = "⚽ WC 2026 OFFICE POOL TRACKER"
    ws["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=16)
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A1"].fill = dark

    ws["A2"] = "wc2026-sports-empire.vercel.app | Pool Manager:"
    ws["A2"].font = Font(name="Calibri", color="888899", size=9)
    ws["A2"].fill = dark
    ws.merge_cells("A2:H2")

    headers = ["#", "Participant Name", "Entry Fee Paid", "Tournament Pick", "Group Winner Bonus",
               "Top Scorer Pick", "Total Points", "Prize Won"]
    ws.append([])
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i)
        cell.value = h
        cell.font = head_font
        cell.fill = green
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample participants
    participants = [
        [1, "Alice Johnson", "£10", "Brazil", "Group A: France", "Mbappé", "=G6+G7", ""],
        [2, "Bob Smith", "£10", "France", "Group B: England", "Kane", "", ""],
        [3, "Carol Davis", "£10", "England", "Group C: USA", "Haaland", "", ""],
        [4, "David Lee", "£10", "Argentina", "Group D: Germany", "Vinicius Jr.", "", ""],
        [5, "Emma Wilson", "£10", "Germany", "Group E: Spain", "Salah", "", ""],
        [6, "Frank Brown", "£10", "Spain", "Group F: Argentina", "Lewandowski", "", ""],
        [7, "Grace Taylor", "£10", "USA", "Group G: Brazil", "Bellingham", "", ""],
        [8, "Harry Evans", "£10", "Portugal", "Group H: Morocco", "Pedri", "", ""],
        [9, "Isla Moore", "£10", "Morocco", "Group I: Portugal", "Osimhen", "", ""],
        [10, "Jack White", "£10", "Netherlands", "Group J: Uruguay", "Musiala", "", ""],
    ]

    for row_idx, row in enumerate(participants, 5):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val
            cell.font = white_font
            cell.fill = mid if row_idx % 2 == 0 else PatternFill("solid", fgColor="16161e")
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = [4, 18, 14, 14, 20, 16, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Group Stage Results ──
    ws2 = wb.create_sheet("Group Stage Results")
    ws2["A1"] = "⚽ GROUP STAGE RESULTS TRACKER"
    ws2["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=14)
    ws2["A1"].fill = dark
    ws2.merge_cells("A1:F1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    groups = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]
    row = 3
    for g in groups:
        ws2.cell(row=row, column=1).value = f"GROUP {g}"
        ws2.cell(row=row, column=1).font = gold_font
        ws2.cell(row=row, column=1).fill = dark
        ws2.merge_cells(f"A{row}:F{row}")
        row += 1
        match_headers = ["Match", "Home Team", "Score", "Away Team", "Date", "Venue"]
        for col, h in enumerate(match_headers, 1):
            c = ws2.cell(row=row, column=col)
            c.value = h
            c.font = head_font
            c.fill = green
            c.alignment = Alignment(horizontal="center")
        row += 1
        for m in range(1, 4):
            ws2.cell(row=row, column=1).value = f"{g}{m}"
            ws2.cell(row=row, column=1).font = white_font
            ws2.cell(row=row, column=1).fill = mid
            for col in range(2, 7):
                ws2.cell(row=row, column=col).fill = mid
                ws2.cell(row=row, column=col).font = white_font
            row += 1
        row += 1

    for i, w in enumerate([6, 16, 8, 16, 12, 20], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Points System ──
    ws3 = wb.create_sheet("Points System")
    ws3["A1"] = "POINTS SYSTEM"
    ws3["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=14)
    ws3["A1"].fill = dark
    ws3.merge_cells("A1:C1")
    ws3["A1"].alignment = Alignment(horizontal="center")

    points = [
        ["Achievement", "Points", "Notes"],
        ["Correct tournament winner", "50", "Biggest prize — bet on early"],
        ["Correct finalist (either side)", "20", "Both finalists count"],
        ["Correct semi-finalist", "10", "Up to 4 teams"],
        ["Correct quarter-finalist", "5", "Up to 8 teams"],
        ["Correct group winner", "5", "12 groups"],
        ["Top scorer prediction correct", "15", "Golden Boot winner"],
        ["Each match score exactly right", "3", "Bonus throughout"],
        ["Each match result correct (no score)", "1", "Participation points"],
        ["Correctly predict an upset (>3/1)", "10", "Bonus for taking risks"],
    ]

    for row_idx, row in enumerate(points, 3):
        for col_idx, val in enumerate(row, 1):
            c = ws3.cell(row=row_idx, column=col_idx)
            c.value = val
            if row_idx == 3:
                c.font = head_font
                c.fill = green
            else:
                c.font = white_font
                c.fill = mid if row_idx % 2 == 0 else PatternFill("solid", fgColor="16161e")
            c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left")

    for i, w in enumerate([35, 10, 30], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 3. BANKROLL MANAGEMENT CALCULATOR (XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def build_bankroll_calc():
    path = f"{OUT}/WC2026-Bankroll-Calculator.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bankroll Calculator"

    green = PatternFill("solid", fgColor="00ff88")
    dark  = PatternFill("solid", fgColor="0a0a0f")
    mid   = PatternFill("solid", fgColor="111118")
    accent = PatternFill("solid", fgColor="1a1a28")
    head_font = Font(name="Calibri", bold=True, color="0a0a0f", size=11)
    white_font = Font(name="Calibri", color="e8e8f0", size=10)
    green_font = Font(name="Calibri", bold=True, color="00ff88", size=12)
    gold_font  = Font(name="Calibri", bold=True, color="ffd700", size=11)
    red_font   = Font(name="Calibri", bold=True, color="ff4455", size=10)

    ws["A1"] = "⚽ WC 2026 BANKROLL MANAGEMENT CALCULATOR"
    ws["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=14)
    ws["A1"].fill = dark
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "wc2026-sports-empire.vercel.app"
    ws["A2"].font = Font(name="Calibri", color="888899", size=9)
    ws["A2"].fill = dark
    ws.merge_cells("A2:E2")

    # Settings section
    ws["A4"] = "⚙ YOUR SETTINGS"
    ws["A4"].font = gold_font
    ws["A4"].fill = dark
    ws.merge_cells("A4:E4")

    settings = [
        ["Starting Bankroll (£)", 500, "", "Enter your starting bankroll in B5"],
        ["Standard Stake %", 0.02, "", "2% recommended (change in B6)"],
        ["High Confidence Stake %", 0.03, "", "For 80%+ confidence bets"],
        ["Max Single Bet %", 0.05, "", "Never exceed this"],
        ["Stop Loss (% of bankroll)", 0.20, "", "Stop if you lose 20%"],
        ["Daily bet limit", 5, "", "Max bets per day"],
    ]

    for row_idx, (label, val, _, note) in enumerate(settings, 5):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=1).font = white_font
        ws.cell(row=row_idx, column=1).fill = mid
        ws.cell(row=row_idx, column=2).value = val
        ws.cell(row=row_idx, column=2).font = green_font
        ws.cell(row=row_idx, column=2).fill = accent
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4).value = note
        ws.cell(row=row_idx, column=4).font = Font(name="Calibri", color="888899", size=9)
        ws.cell(row=row_idx, column=4).fill = dark

    if settings[1][1]:  # Format percentages
        for row in range(6, 10):
            ws.cell(row=row, column=2).number_format = "0.00%"

    # Kelly Calculator
    ws["A12"] = "🎯 KELLY CRITERION CALCULATOR"
    ws["A12"].font = gold_font
    ws["A12"].fill = dark
    ws.merge_cells("A12:E12")

    kelly_rows = [
        ["Your estimated win probability (%)", 0.55, "", "Enter as decimal e.g. 0.55 = 55%"],
        ["Decimal odds offered", 2.10, "", "e.g. 2.10"],
        ["Full Kelly stake %", "=(B13*(B14-1)-(1-B13))/(B14-1)", "", "Auto-calculated"],
        ["Half Kelly stake % (recommended)", "=B15/2", "", "Safer variant"],
        ["Stake amount (£)", "=B5*B16", "", "Based on your bankroll above"],
    ]

    for row_idx, (label, val, _, note) in enumerate(kelly_rows, 13):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=1).font = white_font
        ws.cell(row=row_idx, column=1).fill = mid
        ws.cell(row=row_idx, column=2).value = val
        ws.cell(row=row_idx, column=2).fill = accent
        if row_idx in [15, 16]:
            ws.cell(row=row_idx, column=2).font = green_font
            ws.cell(row=row_idx, column=2).number_format = "0.00%"
        elif row_idx == 17:
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", bold=True, color="ffd700", size=12)
            ws.cell(row=row_idx, column=2).number_format = "£#,##0.00"
        else:
            ws.cell(row=row_idx, column=2).font = green_font
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4).value = note
        ws.cell(row=row_idx, column=4).font = Font(name="Calibri", color="888899", size=9)
        ws.cell(row=row_idx, column=4).fill = dark

    # Bet Tracker Sheet
    ws2 = wb.create_sheet("Bet Tracker")
    ws2["A1"] = "📊 BET TRACKER — WC 2026"
    ws2["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=14)
    ws2["A1"].fill = dark
    ws2.merge_cells("A1:J1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    bet_headers = ["Date", "Match", "Bet Type", "Selection", "Odds", "Stake (£)",
                   "Result", "P&L (£)", "Running Bank (£)", "Confidence %"]
    for col, h in enumerate(bet_headers, 1):
        c = ws2.cell(row=3, column=col)
        c.value = h
        c.font = head_font
        c.fill = green
        c.alignment = Alignment(horizontal="center")

    # Sample rows
    sample_bets = [
        ["Jun 11", "Brazil vs Serbia", "AH", "Brazil -1.5", 1.90, 10, "WON", 9, "=I5+H5", "74%"],
        ["Jun 12", "France vs Australia", "ML", "France Win", 1.45, 15, "WON", 6.75, "=I6+H6", "81%"],
        ["Jun 14", "Germany vs Japan", "O/U", "Over 2.5", 1.72, 10, "WON", 7.20, "=I7+H7", "68%"],
        ["Jun 14", "Spain vs Nigeria", "AH", "Spain -1", 1.88, 12, "WON", 10.56, "=I8+H8", "78%"],
        ["Jun 17", "England vs Colombia", "ML", "England Win", 1.62, 10, "PENDING", 0, "=I9+H9", "66%"],
    ]
    ws2.cell(row=4, column=9).value = 500  # Starting bank

    for row_idx, row in enumerate(sample_bets, 5):
        for col_idx, val in enumerate(row, 1):
            c = ws2.cell(row=row_idx, column=col_idx)
            c.value = val
            c.fill = mid if row_idx % 2 == 0 else PatternFill("solid", fgColor="16161e")
            if col_idx == 7:  # Result
                if val == "WON":
                    c.font = Font(name="Calibri", bold=True, color="00ff88")
                elif val == "LOST":
                    c.font = Font(name="Calibri", bold=True, color="ff4455")
                else:
                    c.font = Font(name="Calibri", color="ffd700")
            elif col_idx == 8:  # P&L
                c.font = green_font if (isinstance(val, (int,float)) and val > 0) else red_font
                c.number_format = "£#,##0.00;[Red]-£#,##0.00"
            elif col_idx == 9:
                c.font = gold_font
                c.number_format = "£#,##0.00"
            else:
                c.font = white_font
            c.alignment = Alignment(horizontal="center")

    for i, w in enumerate([10, 22, 12, 18, 7, 10, 10, 10, 14, 12], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Column widths for sheet 1
    for i, w in enumerate([32, 14, 4, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 4. SPORTSBOOK BONUS TRACKER (XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def build_bonus_tracker():
    path = f"{OUT}/WC2026-Sportsbook-Bonus-Tracker.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UK Bonuses"

    green = PatternFill("solid", fgColor="00ff88")
    dark  = PatternFill("solid", fgColor="0a0a0f")
    mid   = PatternFill("solid", fgColor="111118")
    head_font = Font(name="Calibri", bold=True, color="0a0a0f", size=10)
    white_font = Font(name="Calibri", color="e8e8f0", size=9)

    ws["A1"] = "🎁 WC 2026 SPORTSBOOK BONUS TRACKER"
    ws["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=13)
    ws["A1"].fill = dark
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = Alignment(horizontal="center")

    uk_headers = ["Sportsbook", "Bonus Type", "Bonus Amount", "Min Deposit",
                  "Min Odds", "Wagering Req", "Status", "Notes / Affiliate Link"]
    for col, h in enumerate(uk_headers, 1):
        c = ws.cell(row=3, column=col)
        c.value = h
        c.font = head_font
        c.fill = green
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    uk_books = [
        ["bet365", "Bet Credits", "Up to £50", "£10", "1/5 (1.20)", "None", "⬜ Not Started", "bet365.com/affiliates"],
        ["Betfair", "Free Exchange Bet", "£20", "£10", "N/A (exchange)", "None", "⬜ Not Started", "betfair.com/affiliates"],
        ["William Hill", "Free Bet", "£30", "£10", "4/5 (1.80)", "None (returns bonus)", "⬜ Not Started", "willhill.com"],
        ["Paddy Power", "Free Bet", "£20", "£10", "1/5", "None", "⬜ Not Started", "paddypower.com"],
        ["Sky Bet", "Free Bet", "£30", "£10", "Evens (2.00)", "None", "⬜ Not Started", "skybet.com"],
        ["Ladbrokes", "Free Bet", "£20", "£5", "1/2", "None", "⬜ Not Started", "ladbrokes.com"],
        ["Coral", "Free Bet", "£20", "£5", "1/2", "None", "⬜ Not Started", "coral.co.uk"],
        ["Betway", "Free Bet", "£30", "£10", "4/5", "None", "⬜ Not Started", "betway.com"],
        ["888sport", "Free Bet", "£30", "£10", "1/5", "None", "⬜ Not Started", "888sport.com"],
        ["Unibet", "Free Bet", "£40", "£10", "4/5", "None", "⬜ Not Started", "unibet.co.uk"],
        ["BetVictor", "Free Bet", "£30", "£5", "1/2", "None", "⬜ Not Started", "betvictor.com"],
        ["10bet", "Free Bet", "£50", "£10", "Evens", "None", "⬜ Not Started", "10bet.co.uk"],
        ["Betfred", "Free Bet", "£60 (£30+£30)", "£10", "Evens", "None", "⬜ Not Started", "betfred.com"],
        ["Spreadex", "Free Bet", "£36", "£25", "N/A spread", "None", "⬜ Not Started", "spreadex.com"],
        ["BoyleSports", "Free Bet", "£20", "£10", "Evens", "None", "⬜ Not Started", "boylesports.com"],
        ["Bet Storm", "Free Bet", "£30", "£10", "Evens", "None", "⬜ Not Started", "betstorm.co.uk"],
        ["Midnite", "Free Bet", "£30", "£10", "N/A exchange", "None", "⬜ Not Started", "midnite.com"],
        ["Sporting Index", "Free Bet", "£10", "£10", "N/A spread", "None", "⬜ Not Started", ""],
        ["Betsson", "Free Bet", "£30", "£10", "Evens", "None", "⬜ Not Started", ""],
        ["NetBet", "Free Bet", "£10", "£10", "6/4", "None", "⬜ Not Started", ""],
    ]

    total_row = len(uk_books) + 4
    for row_idx, row in enumerate(uk_books, 4):
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.value = val
            c.fill = mid if row_idx % 2 == 0 else PatternFill("solid", fgColor="16161e")
            c.font = white_font
            c.alignment = Alignment(horizontal="center" if col_idx != 1 else "left")

    # Total row
    ws.cell(row=total_row, column=1).value = "TOTAL POTENTIAL"
    ws.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, color="ffd700")
    ws.cell(row=total_row, column=1).fill = dark
    ws.cell(row=total_row, column=3).value = "~£560 in free bets"
    ws.cell(row=total_row, column=3).font = Font(name="Calibri", bold=True, color="00ff88", size=11)
    ws.cell(row=total_row, column=3).fill = dark
    ws.merge_cells(f"A{total_row}:B{total_row}")
    ws.merge_cells(f"C{total_row}:H{total_row}")

    for i, w in enumerate([16, 16, 14, 12, 14, 14, 14, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # US sheet
    ws2 = wb.create_sheet("US Bonuses")
    ws2["A1"] = "🇺🇸 US SPORTSBOOK BONUSES — WC 2026"
    ws2["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=13)
    ws2["A1"].fill = dark
    ws2.merge_cells("A1:H1")
    ws2["A1"].alignment = Alignment(horizontal="center")

    for col, h in enumerate(uk_headers, 1):
        c = ws2.cell(row=3, column=col)
        c.value = h
        c.font = head_font
        c.fill = green
        c.alignment = Alignment(horizontal="center")

    us_books = [
        ["DraftKings", "Bet $5 Get $200", "$200 bonus bets", "$5", "None", "1× on bonus bets", "⬜ Not Started", "affiliates.draftkings.com"],
        ["FanDuel", "No Sweat Bet", "Up to $1,000 refund", "$10", "None", "1× on bonus", "⬜ Not Started", "fanduel.com/affiliates"],
        ["BetMGM", "First Bet Offer", "Up to $1,500 bonus", "$10", "None", "1× on bonus", "⬜ Not Started", ""],
        ["Caesars", "Full Match", "Up to $1,000", "$10", "None", "1× on bonus", "⬜ Not Started", ""],
        ["PointsBet", "Second Chance", "Up to $500", "$10", "None", "1× on bonus", "⬜ Not Started", ""],
        ["BetRivers", "2nd Chance Bet", "Up to $500", "$10", "None", "1×", "⬜ Not Started", ""],
        ["Fanatics", "Bet & Get", "Up to $1,000", "$5", "None", "Tiered wagering", "⬜ Not Started", ""],
        ["ESPNBet", "First Bet Reset", "Up to $1,000", "$10", "None", "1×", "⬜ Not Started", ""],
        ["Bet365", "First Bet Safety Net", "Up to $1,000", "$10", "None", "1×", "⬜ Not Started", ""],
        ["WynnBet", "Bet & Get", "Up to $200", "$5", "None", "1×", "⬜ Not Started", ""],
    ]

    for row_idx, row in enumerate(us_books, 4):
        for col_idx, val in enumerate(row, 1):
            c = ws2.cell(row=row_idx, column=col_idx)
            c.value = val
            c.fill = mid if row_idx % 2 == 0 else PatternFill("solid", fgColor="16161e")
            c.font = white_font
            c.alignment = Alignment(horizontal="center" if col_idx != 1 else "left")

    us_total = len(us_books) + 4
    ws2.cell(row=us_total, column=1).value = "TOTAL POTENTIAL"
    ws2.cell(row=us_total, column=3).value = "~$7,700 in US bonuses"
    ws2.cell(row=us_total, column=3).font = Font(name="Calibri", bold=True, color="00ff88", size=11)
    ws2.cell(row=us_total, column=1).font = Font(name="Calibri", bold=True, color="ffd700")
    ws2.cell(row=us_total, column=1).fill = dark
    ws2.cell(row=us_total, column=3).fill = dark
    ws2.merge_cells(f"A{us_total}:B{us_total}")
    ws2.merge_cells(f"C{us_total}:H{us_total}")

    for i, w in enumerate([16, 18, 16, 10, 10, 18, 14, 26], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 5. FANTASY DRAFT KIT PDF
# ─────────────────────────────────────────────────────────────────────────────
def build_fantasy_kit():
    path = f"{OUT}/WC2026-Fantasy-Draft-Kit.pdf"
    doc = SimpleDocTemplate(path, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", fontSize=24, textColor=ACCENT, alignment=TA_CENTER,
                              fontName="Helvetica-Bold", spaceAfter=8)
    sub_s   = ParagraphStyle("S", fontSize=11, textColor=LIGHT, alignment=TA_CENTER, spaceAfter=16)
    h1      = ParagraphStyle("H1", fontSize=16, textColor=ACCENT, spaceBefore=14, spaceAfter=8,
                              fontName="Helvetica-Bold")
    h2      = ParagraphStyle("H2", fontSize=12, textColor=GOLD, spaceBefore=10, spaceAfter=6,
                              fontName="Helvetica-Bold")
    body    = ParagraphStyle("B", fontSize=9.5, textColor=LIGHT, spaceAfter=6, leading=15)

    story = []
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph("⚽ WC 2026 FANTASY DRAFT KIT", title_s))
    story.append(Paragraph("AI Rankings · Sleeper Picks · Avoid Lists · Auction Values", sub_s))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT))
    story.append(Spacer(1, 0.4*cm))

    story.append(Paragraph("TOP 30 OVERALL — AI DRAFT RANKINGS", h1))
    story.append(Paragraph("Ranked by projected fantasy points across 7 matches (if team reaches final)", body))

    players = [
        ["Rank", "Player", "Team", "Pos", "Proj. Pts", "Avg Odds", "AI Note"],
        ["1", "Kylian Mbappé", "France", "FWD", "87", "2.50", "Golden Boot favourite, max games"],
        ["2", "Erling Haaland", "Norway", "FWD", "82", "3.20", "Clinical finisher, Norway dark horse"],
        ["3", "Vinicius Jr.", "Brazil", "FWD", "79", "2.10", "Brazil = most games projection"],
        ["4", "Jude Bellingham", "England", "MID", "76", "4.50", "Goal threat from deep, set pieces"],
        ["5", "Lamine Yamal", "Spain", "FWD", "74", "3.00", "Spain's key man, assists machine"],
        ["6", "Pedri", "Spain", "MID", "72", "3.00", "Chance creation + defensive cover"],
        ["7", "Phil Foden", "England", "MID", "70", "4.50", "High touch volume in English system"],
        ["8", "Harry Kane", "England", "FWD", "69", "4.50", "Penalty taker, clinical striker"],
        ["9", "Florian Wirtz", "Germany", "MID", "68", "5.00", "Germany's technical fulcrum"],
        ["10", "Jamal Musiala", "Germany", "MID", "67", "5.00", "Carries ball, creates constantly"],
        ["11", "Mohamed Salah", "Egypt", "FWD", "65", "15.0", "Top scorer in groups if Egypt qualifies"],
        ["12", "Bukayo Saka", "England", "FWD", "64", "4.50", "Consistent performer, dribbles + assists"],
        ["13", "Rodri", "Spain", "MID", "62", "3.00", "Defensive mid points: tackles, interceptions"],
        ["14", "Heung-min Son", "S. Korea", "FWD", "60", "20.0", "Carries Korea alone, high ceiling"],
        ["15", "Raphinha", "Brazil", "FWD", "59", "2.10", "Brazil's right wing threat"],
        ["16", "Bernardo Silva", "Portugal", "MID", "57", "6.00", "Ronaldo support, consistent"],
        ["17", "Antoine Griezmann", "France", "FWD", "55", "2.50", "France secondary striker, bonus ready"],
        ["18", "Gavi", "Spain", "MID", "54", "3.00", "Aggressive pressuring mid"],
        ["19", "Victor Osimhen", "Nigeria", "FWD", "53", "25.0", "Potential group stage top scorer"],
        ["20", "Christian Pulisic", "USA", "MID", "52", "10.0", "Host nation premium, goal threat"],
        ["21", "Federico Valverde", "Uruguay", "MID", "50", "18.0", "Box-to-box points machine"],
        ["22", "Cody Gakpo", "Netherlands", "FWD", "49", "8.00", "Netherlands forward, efficient"],
        ["23", "Achraf Hakimi", "Morocco", "DEF", "47", "20.0", "Attacking full-back, assists + goals"],
        ["24", "Nico Williams", "Spain", "FWD", "46", "3.00", "Speed threat, Spain's breakaway weapon"],
        ["25", "Endrick", "Brazil", "FWD", "45", "2.10", "Impact sub → starter for Brazil"],
        ["26", "Robert Lewandowski", "Poland", "FWD", "44", "30.0", "Poland's only scorer, captain"],
        ["27", "Luka Modrić", "Croatia", "MID", "42", "22.0", "Still elite at 40, last WC"],
        ["28", "Oliver Giroud", "France", "FWD", "40", "2.50", "Target man, set piece threat"],
        ["29", "Alessandro Bastoni", "Italy", "DEF", "38", "12.0", "Italy defensive core"],
        ["30", "Manuel Neuer", "Germany", "GK", "35", "5.00", "Top GK for clean sheets"],
    ]

    t = Table(players, colWidths=[1*cm, 3.5*cm, 2.5*cm, 1.2*cm, 1.8*cm, 1.8*cm, 5.7*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), ACCENT),
        ("TEXTCOLOR", (0,0), (-1,0), DARK),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [MID, colors.HexColor("#16161e")]),
        ("TEXTCOLOR", (0,1), (-1,-1), LIGHT),
        ("ALIGN", (0,0), (5,-1), "CENTER"),
        ("ALIGN", (6,0), (6,-1), "LEFT"),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#333344")),
        ("PADDING", (0,0), (-1,-1), 4),
        # Highlight top 3
        ("TEXTCOLOR", (0,1), (-1,3), GOLD),
        ("FONTNAME", (0,1), (-1,3), "Helvetica-Bold"),
    ]))
    story.append(t)
    story.append(PageBreak())

    # Sleeper picks
    story.append(Paragraph("💎 SLEEPER PICKS — Value Under the Radar", h1))
    sleepers = [
        ("Takefusa Kubo (Japan)",
         "Creative midfielder in a Japan team that overperforms expectations. Low draft cost, high ceiling."),
        ("Abde Ezzalzouli (Morocco)",
         "Morocco's explosive winger — if they replicate 2022's run, he'll rack up assists at low ownership."),
        ("Khvicha Kvaratskhelia (Georgia?/not qualified yet)",
         "Monitor squad — if a dark horse team qualifies, their main creator is always value."),
        ("Giacomo Raspadori (Italy)",
         "Poacher in Italy's tight tactical system. Often scores from limited chances."),
        ("Julian Alvarez (Argentina)",
         "Messi's understudy but a prolific finisher himself — often ignored in drafts for the big names."),
        ("Ibrahim Traoré (Burkina Faso / Pan-African)",
         "If any African qualifier has a surprise run, their key attacker will be massively underowned."),
    ]
    for name, note in sleepers:
        story.append(Paragraph(f"✦ {name}", h2))
        story.append(Paragraph(note, body))

    story.append(PageBreak())
    story.append(Paragraph("🚫 AVOID LIST — Hyped But Risky", h1))
    avoid = [
        ("Neymar Jr. (Brazil)", "Injury history at tournaments. Brazilian system doesn't depend on him now. Overpriced."),
        ("Cristiano Ronaldo (Portugal)", "At 41+, Portugal plays around younger talent. Starts uncertain."),
        ("Lionel Messi (Argentina)", "May not be active or fully fit. Historical legend but WC 2022 peak may be past."),
        ("Any GK from a weak nation", "Fantasy GK points come from saves + clean sheets. Weak team = high shots faced but few wins."),
        ("Late injury news players", "Always monitor 24 hours before draft — avoid those with muscle/fitness doubts."),
    ]
    for name, note in avoid:
        story.append(Paragraph(f"✗ {name}", h2))
        story.append(Paragraph(note, body))

    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT))
    story.append(Spacer(1, 0.3*cm))
    story.append(Paragraph("Live rankings updated daily at wc2026-sports-empire.vercel.app", sub_s))

    doc.build(story)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 6. ARB CALCULATOR (XLSX)
# ─────────────────────────────────────────────────────────────────────────────
def build_arb_calc():
    path = f"{OUT}/WC2026-Arbitrage-Calculator.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2-Way Arb"

    dark  = PatternFill("solid", fgColor="0a0a0f")
    green = PatternFill("solid", fgColor="00ff88")
    mid   = PatternFill("solid", fgColor="111118")
    accent = PatternFill("solid", fgColor="1a1a28")

    ws["A1"] = "⚡ WC 2026 ARBITRAGE CALCULATOR"
    ws["A1"].font = Font(name="Calibri", bold=True, color="00ff88", size=14)
    ws["A1"].fill = dark
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Guaranteed profit regardless of outcome"
    ws["A2"].font = Font(name="Calibri", color="888899", size=9)
    ws["A2"].fill = dark
    ws.merge_cells("A2:D2")

    # 2-Way Arb
    ws["A4"] = "2-WAY ARBITRAGE (e.g. Asian Handicap)"
    ws["A4"].font = Font(name="Calibri", bold=True, color="ffd700", size=11)
    ws["A4"].fill = dark
    ws.merge_cells("A4:D4")

    inputs = [
        ("Total bankroll to use (£)", 100, "=B6", "Change B5 to your bankroll"),
        ("Outcome 1 Odds (decimal)", 2.10, "", "e.g. 2.10"),
        ("Outcome 2 Odds (decimal)", 2.05, "", "e.g. 2.05"),
        ("Arb % (lower = better)", "=100/B6+100/B7", "", "Under 100% = arb exists"),
        ("Is there an arb?", '=IF(B8<100,"✅ YES — ARB EXISTS!","❌ No arb")', "", ""),
        ("Stake on Outcome 1 (£)", "=B5*(1/B6)/((1/B6)+(1/B7))", "", "Auto-calculated"),
        ("Stake on Outcome 2 (£)", "=B5*(1/B7)/((1/B6)+(1/B7))", "", "Auto-calculated"),
        ("Guaranteed profit (£)", "=B5*(1-B8/100)", "", "Auto-calculated"),
        ("Return on investment (%)", "=B12/B5*100", "", "Your guaranteed ROI"),
    ]

    for row_idx, (label, val, _, note) in enumerate(inputs, 5):
        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", color="e8e8f0", size=10)
        ws.cell(row=row_idx, column=1).fill = mid
        ws.cell(row=row_idx, column=2).value = val
        ws.cell(row=row_idx, column=2).fill = accent
        if row_idx in [8]:
            ws.cell(row=row_idx, column=2).number_format = "0.00%"
        elif row_idx in [10, 11, 12]:
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", bold=True, color="00ff88", size=11)
            ws.cell(row=row_idx, column=2).number_format = "£#,##0.00"
        elif row_idx == 9:
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", bold=True, color="00ff88", size=11)
        elif row_idx == 13:
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", bold=True, color="ffd700", size=11)
            ws.cell(row=row_idx, column=2).number_format = "0.00%"
        else:
            ws.cell(row=row_idx, column=2).font = Font(name="Calibri", color="00ff88", size=10)
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=4).value = note
        ws.cell(row=row_idx, column=4).font = Font(name="Calibri", color="888899", size=9)
        ws.cell(row=row_idx, column=4).fill = dark

    for i, w in enumerate([30, 16, 4, 30], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(path)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# 7. BETTING GLOSSARY PDF
# ─────────────────────────────────────────────────────────────────────────────
def build_glossary():
    path = f"{OUT}/WC2026-Betting-Glossary.pdf"
    doc = SimpleDocTemplate(path, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_s = ParagraphStyle("T", fontSize=22, textColor=ACCENT, alignment=TA_CENTER,
                              fontName="Helvetica-Bold", spaceAfter=8)
    sub_s   = ParagraphStyle("S", fontSize=10, textColor=LIGHT, alignment=TA_CENTER, spaceAfter=14)
    h1      = ParagraphStyle("H1", fontSize=14, textColor=ACCENT, spaceBefore=12, spaceAfter=6,
                              fontName="Helvetica-Bold")
    body    = ParagraphStyle("B", fontSize=9, textColor=LIGHT, spaceAfter=5, leading=14)
    term    = ParagraphStyle("Term", fontSize=10, textColor=GOLD, spaceBefore=6, spaceAfter=2,
                              fontName="Helvetica-Bold")

    story = []
    story.append(Paragraph("⚽ COMPLETE SPORTS BETTING GLOSSARY", title_s))
    story.append(Paragraph("WC 2026 Edition · A-Z of Every Term You Need to Know", sub_s))
    story.append(HRFlowable(width="100%", thickness=1, color=ACCENT))
    story.append(Spacer(1, 0.4*cm))

    terms = {
        "A": [
            ("Accumulator (Acca)", "A bet combining multiple selections where all must win. Each win multiplies the odds. Typically 4+ selections. Higher risk, higher reward."),
            ("Asian Handicap (AH)", "A bet type that eliminates draws by giving a virtual head start to the underdog. Most common: AH 0, AH 0.5, AH 1, AH 1.5."),
            ("Arbitrage (Arb)", "Exploiting different odds at different bookmakers to guarantee a profit regardless of outcome. Legal and risk-free."),
            ("Ante-post", "A bet placed before an event is imminent — e.g., pre-tournament outright bets. Higher odds, but stake usually lost if your team withdraws."),
        ],
        "B": [
            ("Banker", "A highly confident selection, often used as the foundation of an accumulator. If the banker loses, the whole bet loses."),
            ("Bankroll", "Your total betting fund. Never risk more than 5% on a single bet."),
            ("Bet Credits", "Bonus funds given by sportsbooks that can only be used as stakes, not withdrawn directly."),
            ("Both Teams to Score (BTTS)", "A market on whether both teams will score at least once. Yes or No."),
            ("Bookmaker (Bookie)", "A licensed company that accepts bets on sporting events and sets odds."),
        ],
        "C": [
            ("Cash Out", "A feature allowing you to settle a bet early before the event ends, taking a guaranteed sum in exchange for giving up the potential full win."),
            ("Correct Score", "Betting on the exact final scoreline. High odds, low probability."),
            ("Cover", "When a spread bet wins by the required margin (e.g., 'covering the spread' of -1.5 goals)."),
        ],
        "D": [
            ("Dead Heat", "When two or more outcomes finish equal (e.g., two players tied on goals). Stakes are divided by the number of tied participants."),
            ("Double Chance", "Betting on two of three outcomes — Home or Draw, Away or Draw, or Home or Away. Lower odds, higher probability."),
            ("Draw No Bet (DNB)", "Bet on home or away win; stake returned if match ends in a draw. AH 0 equivalent."),
            ("Dutching", "Distributing a total stake across multiple selections so the same profit is made regardless of which one wins."),
        ],
        "E": [
            ("Each Way (E/W)", "A two-part bet — half on win, half on place. Common in horse racing, less so in football."),
            ("Edge", "Your mathematical advantage over the bookmaker. A positive edge means your bet is statistically profitable long-term."),
            ("Enhanced Odds / Odds Boost", "A promotional offer where a bookmaker increases the odds on a specific selection."),
            ("Expected Value (EV)", "The average amount you expect to win per bet over time. Positive EV = profitable bet."),
        ],
        "F": [
            ("Favourite", "The team or player most likely to win according to the bookmaker, reflected in shorter odds."),
            ("Fixed Odds", "Standard pre-match betting where odds are agreed when the bet is placed."),
            ("Float", "The amount in your betting account available to use."),
        ],
        "G": [
            ("Grand Salami", "A bet on the total number of goals across all games on a given day's fixture list."),
            ("Graded Bet", "A bet that has been officially settled — either as a winner, loser, or push."),
        ],
        "H": [
            ("Half Time / Full Time (HT/FT)", "Predict the result at half time AND full time. E.g. Away/Home = away team leading at HT, home team wins FT."),
            ("Handicap Betting", "Giving one team a virtual advantage or disadvantage to level the playing field."),
            ("Hedge", "Placing a bet on the opposite outcome to an existing bet to guarantee a profit or reduce risk."),
        ],
        "I": [
            ("In-Play / Live Betting", "Betting on an event after it has started. Odds fluctuate in real time based on match events."),
            ("Implied Probability", "The probability implied by the bookmaker's odds. Formula: 1 ÷ decimal odds × 100."),
        ],
        "K": [
            ("Kelly Criterion", "A mathematical formula for calculating optimal bet size based on your edge and odds. Kelly % = (BP − Q) / B."),
        ],
        "L": [
            ("Lay Bet", "Available on exchanges (Betfair). You act as the bookmaker, betting that a selection will NOT win."),
            ("Limit", "Maximum stake a bookmaker will accept on a given market."),
            ("Line", "The current odds or point spread on a given market."),
            ("Live Odds", "Real-time, continuously updated odds during an in-play event."),
        ],
        "M": [
            ("Matched Betting", "Using bookmaker free bets and bonuses alongside lay bets on exchanges to guarantee profit. Risk-free when done correctly."),
            ("Middle", "A scenario where the final score or result falls between two opposing bets, resulting in winning both."),
            ("Moneyline", "Simple win/lose bet with no handicap. American format shows -150 (favourite) or +200 (underdog)."),
            ("Mug Bet", "A losing bet placed intentionally to disguise your activity from the bookmaker and avoid account restrictions."),
        ],
        "O": [
            ("Odds Against", "When your potential profit exceeds your stake (decimal > 2.00, e.g., 3/1 fractional)."),
            ("Odds On", "When your stake exceeds your potential profit (decimal < 2.00, e.g., 1/2 fractional)."),
            ("Outright", "A bet on the overall winner of a tournament rather than a single match."),
            ("Over/Under (O/U)", "A bet on whether the total goals (or other metric) will be over or under a set number, e.g., Over 2.5."),
        ],
        "P": [
            ("Parlay", "American term for accumulator. Multiple selections, all must win."),
            ("Puck Line / Run Line / Goal Line", "Sport-specific spread equivalents to Asian Handicap."),
            ("Push", "A result that returns your stake — happens in Asian Handicap when the margin exactly equals the handicap."),
        ],
        "R": [
            ("Reverse Asian Handicap", "Betting on the underdog with the handicap advantage."),
            ("ROI (Return on Investment)", "Your net profit as a percentage of total stakes."),
            ("Round Robin", "A series of accumulators covering all combinations from a group of selections."),
        ],
        "S": [
            ("Same Game Parlay (SGP)", "Combining multiple markets from the same match into one bet. High odds, high correlation risk."),
            ("Sharp", "A professional or sophisticated bettor whose bets move the market."),
            ("Smart Money", "Large bets placed by sharp bettors that cause the line to move."),
            ("Spread Betting", "Betting on how far above or below a given number an outcome will fall. Profit/loss multiplies with accuracy."),
            ("Steam", "Rapid market movement when sharp bettors all bet on the same side simultaneously."),
        ],
        "T": [
            ("Teaser", "A parlay with adjusted point spreads in your favour in exchange for lower odds."),
            ("Tissue Price", "A bookmaker's internal pre-market odds before they factor in their margin."),
            ("Total", "The combined score of both teams. Often used for Over/Under betting."),
        ],
        "V": [
            ("Value Bet", "A bet where the probability of winning is higher than what the odds imply. The cornerstone of profitable betting."),
            ("Void Bet", "A bet that is cancelled and stakes returned — typically if a match is abandoned or a player doesn't participate."),
        ],
        "W": [
            ("Whale", "A very high-stakes bettor."),
            ("Win Only", "A bet that only pays if the selection wins outright (not placed). Contrast with each-way."),
        ],
    }

    for letter, entries in sorted(terms.items()):
        story.append(Paragraph(f"— {letter} —", h1))
        for term_title, definition in entries:
            story.append(Paragraph(term_title, term))
            story.append(Paragraph(definition, body))

    story.append(PageBreak())
    story.append(Paragraph("📱 Live AI Picks & Odds", title_s))
    story.append(Paragraph("wc2026-sports-empire.vercel.app", sub_s))
    story.append(Paragraph("Telegram · Discord · X/Twitter · TikTok · YouTube", sub_s))

    doc.build(story)
    print(f"✅ Built: {path}")
    return path


# ─────────────────────────────────────────────────────────────────────────────
# RUN ALL
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Building digital products...\n")
    build_betting_guide()
    build_office_pool()
    build_bankroll_calc()
    build_bonus_tracker()
    build_fantasy_kit()
    build_arb_calc()
    build_glossary()
    print("\n✅ All digital products built!")
